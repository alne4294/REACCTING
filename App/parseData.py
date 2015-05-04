#! /usr/bin/python


import openpyxl as px
import json

# only used if run directly. (Not used as a library)
from sys import argv

def parseWorkbook(filePath):
    workbook = px.load_workbook(filePath,  guess_types=True)
    worksheet = workbook["AllSamplesTogether"]
    data = {}
    for row in map(lambda i: i+1, range(worksheet.get_highest_row())):
        if None != worksheet.cell(row=row, column=1).value:
            data[worksheet.cell(row=row, column=8).value.isoformat()] = {
                'location': worksheet.cell(row=row, column=9).value,
                'stove': worksheet.cell(row=row, column=1).value,
                'cookingTimes': worksheet.cell(row=row, column=11).value,
                'massFuelUsed': worksheet.cell(row=row, column=10).value,
                'COEF': worksheet.cell(row=row, column=2).value,
                'CO2EF': worksheet.cell(row=row, column=3).value,
                'NOEF': worksheet.cell(row=row, column=4).value,
                'NO2EF': worksheet.cell(row=row, column=5).value,
                'VOCEF': worksheet.cell(row=row, column=6).value,
                'MCE': worksheet.cell(row=row, column=7).value
            }
    return data

if __name__ == "__main__":
    data = parseWorkbook(argv[1])
    outFile = open('datafile.json', 'w')
    outFile.write(json.dumps(data))
    outFile.close()
